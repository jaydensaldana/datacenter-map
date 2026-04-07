"""
fetch_data.py — Automated data pipeline for USA Data Center Map
================================================================
Reads manual data from Google Sheet, fetches federal data, merges,
and writes data/combined.json for the map builder.

DESIGN: All data keyed by state name. Rows can be added/reordered
in the Google Sheet without breaking anything. Month columns are
identified by header text (e.g., "Apr 2026").

SOURCES:
  - Google Sheet:  DC counts, tax incentives, regulations, regional markets
  - EIA EPM 5.6.A: Average price of electricity by state/sector
  - BLS QCEW:      Employment by NAICS (518210, 5415, 236220)
  - Census PEP:    State population estimates
"""

import requests, csv, io, json, time, os, re
from datetime import datetime

GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID", "1ZtNh_GJCEzZ3zNxIjJad09ihbwT5UDBPfy7wK2f1gWo")
CURRENT_YEAR = datetime.now().year
CHANGE_WINDOWS = [1, 3, 5]
MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
MONTH_RE = re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+(\d{4})$')

STATE_ABBREVS = {
    'Alabama':'AL','Alaska':'AK','Arizona':'AZ','Arkansas':'AR','California':'CA',
    'Colorado':'CO','Connecticut':'CT','Delaware':'DE','Florida':'FL','Georgia':'GA',
    'Hawaii':'HI','Idaho':'ID','Illinois':'IL','Indiana':'IN','Iowa':'IA','Kansas':'KS',
    'Kentucky':'KY','Louisiana':'LA','Maine':'ME','Maryland':'MD','Massachusetts':'MA',
    'Michigan':'MI','Minnesota':'MN','Mississippi':'MS','Missouri':'MO','Montana':'MT',
    'Nebraska':'NE','Nevada':'NV','New Hampshire':'NH','New Jersey':'NJ',
    'New Mexico':'NM','New York':'NY','North Carolina':'NC','North Dakota':'ND',
    'Ohio':'OH','Oklahoma':'OK','Oregon':'OR','Pennsylvania':'PA','Rhode Island':'RI',
    'South Carolina':'SC','South Dakota':'SD','Tennessee':'TN','Texas':'TX','Utah':'UT',
    'Vermont':'VT','Virginia':'VA','Washington':'WA','West Virginia':'WV',
    'Wisconsin':'WI','Wyoming':'WY'
}
ABBREV_TO_STATE = {v: k for k, v in STATE_ABBREVS.items()}
STATE_FIPS = {
    'AL':'01','AK':'02','AZ':'04','AR':'05','CA':'06','CO':'08','CT':'09','DE':'10',
    'FL':'12','GA':'13','HI':'15','ID':'16','IL':'17','IN':'18','IA':'19','KS':'20',
    'KY':'21','LA':'22','ME':'23','MD':'24','MA':'25','MI':'26','MN':'27','MS':'28',
    'MO':'29','MT':'30','NE':'31','NV':'32','NH':'33','NJ':'34','NM':'35','NY':'36',
    'NC':'37','ND':'38','OH':'39','OK':'40','OR':'41','PA':'42','RI':'44','SC':'45',
    'SD':'46','TN':'47','TX':'48','UT':'49','VT':'50','VA':'51','WA':'53','WV':'54',
    'WI':'55','WY':'56'
}
FIPS_TO_ABBREV = {v: k for k, v in STATE_FIPS.items()}


def safe_int(v):
    if v is None or v == '': return None
    try: return int(float(str(v).replace(',','').strip()))
    except: return None

def safe_float(v):
    if v is None or v in ('', '--', 'N/A', 'NA'): return None
    try: return float(str(v).replace(',','').replace('$','').strip())
    except: return None

def parse_month(label):
    m = MONTH_RE.match(label.strip())
    if not m: return None
    return (int(m.group(2)), MONTH_NAMES.index(m.group(1)) + 1)

def pct_change(cur, hist):
    if cur is None or hist is None or hist == 0: return None
    return (cur - hist) / hist


# ═══════════════════════════════════════════
# GOOGLE SHEETS
# ═══════════════════════════════════════════
def fetch_sheet_xlsx_fallback(name):
    """
    Fallback: download the entire Google Sheet as XLSX and parse the named tab.
    This avoids CSV multi-line cell issues entirely.
    """
    print(f"  XLSX fallback for '{name}'...")
    import openpyxl
    url = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=xlsx"
    try:
        r = requests.get(url, timeout=60)
        if r.status_code != 200:
            print(f"    Status {r.status_code}")
            return [], []
        with open('data/gsheet_dump.xlsx', 'wb') as f:
            f.write(r.content)
        wb = openpyxl.load_workbook('data/gsheet_dump.xlsx', data_only=True)
        if name not in wb.sheetnames:
            print(f"    Tab '{name}' not in {wb.sheetnames}")
            return [], []
        ws = wb[name]
        # First row = headers
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            headers.append(str(v).strip() if v else '')
        # Subsequent rows = data
        rows = []
        for r_idx in range(2, ws.max_row + 1):
            first_val = ws.cell(row=r_idx, column=1).value
            if not first_val: continue
            row_dict = {}
            for c_idx, h in enumerate(headers, 1):
                if not h: continue
                v = ws.cell(row=r_idx, column=c_idx).value
                row_dict[h] = str(v).strip() if v is not None else ''
            rows.append(row_dict)
        print(f"    XLSX fallback: {len(rows)} rows, {len(headers)} cols")
        return rows, headers
    except Exception as e:
        print(f"    XLSX fallback error: {e}")
        return [], []

def fetch_sheet(name, debug=False):
    enc = requests.utils.quote(name)
    # Try multiple URL patterns. The export?format=csv endpoint sometimes
    # handles multi-line cells better than gviz.
    urls = [
        f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/gviz/tq?tqx=out:csv&sheet={enc}",
        f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=csv&sheet={enc}",
    ]
    for url_idx, url in enumerate(urls):
        try:
            r = requests.get(url, timeout=30)
            if r.status_code != 200 or len(r.text) < 10:
                continue

            text = r.text

            if debug:
                print(f"  DEBUG: URL {url_idx}: {len(text)} bytes")
                print(f"  DEBUG: First 300 chars: {text[:300]!r}")
                print(f"  DEBUG: Number of newlines: {text.count(chr(10))}")
                print(f"  DEBUG: Number of carriage returns: {text.count(chr(13))}")

            # Use csv.reader on the full text — it correctly handles
            # quoted fields containing embedded newlines (e.g., long descriptions)
            reader = csv.reader(io.StringIO(text), quotechar='"', skipinitialspace=True)
            all_rows = list(reader)
            if not all_rows:
                continue

            headers = [h.strip().strip('"') for h in all_rows[0]]
            rows = []
            for row in all_rows[1:]:
                if len(row) >= 1 and row[0].strip():
                    rows.append({h: (row[i].strip() if i < len(row) else '') for i, h in enumerate(headers)})

            print(f"  '{name}': {len(rows)} rows, {len(headers)} cols (URL {url_idx})")

            # If we got too few rows, try XLSX fallback
            if len(rows) < 5:
                print(f"    Too few rows from CSV, trying XLSX fallback...")
                xlsx_rows, xlsx_headers = fetch_sheet_xlsx_fallback(name)
                if len(xlsx_rows) >= 5:
                    return xlsx_rows, xlsx_headers
                # Continue to next URL if XLSX also fails
                continue

            return rows, headers
        except Exception as e:
            print(f"  Warning ({url_idx}): {e}")
            continue

    # Last resort: XLSX fallback
    print(f"  All CSV URLs failed for '{name}', trying XLSX fallback")
    return fetch_sheet_xlsx_fallback(name)

def read_dc_history(rows, headers):
    month_cols = {}
    for h in headers:
        p = parse_month(h)
        if p: month_cols[p] = h
    history, latest = {}, (2020, 1)
    for row in rows:
        state = row.get('State', '').strip()
        if state not in STATE_ABBREVS: continue
        history[state] = {}
        for ym, hdr in month_cols.items():
            v = safe_int(row.get(hdr, ''))
            if v is not None:
                history[state][ym] = v
                if ym > latest: latest = ym
    print(f"  DC History: {len(history)} states, latest={MONTH_NAMES[latest[1]-1]} {latest[0]}")
    return history, latest

def read_markets(rows, headers):
    month_cols = {}
    for h in headers:
        p = parse_month(h)
        if p: month_cols[p] = h
    markets, latest = {}, (2020, 1)
    for row in rows:
        state = row.get('State', '').strip()
        city = row.get('Market/City', '').strip()
        if not state or not city: continue
        markets[(state, city)] = {}
        for ym, hdr in month_cols.items():
            v = safe_int(row.get(hdr, ''))
            if v is not None:
                markets[(state, city)][ym] = v
                if ym > latest: latest = ym
    print(f"  Markets: {len(markets)} entries")
    return markets, latest

def read_tax(rows, headers):
    """
    Robustly map Tax Incentive headers regardless of exact wording.
    Google Sheets CSV export sometimes alters special characters like '?'.
    """
    # Find the actual header keys by partial-match
    header_map = {}
    for h in headers:
        h_lower = h.lower().strip()
        if 'tax incentive' in h_lower:
            header_map['tax_incentive'] = h
        elif 'incentive type' in h_lower or h_lower == 'type':
            header_map['incentive_type'] = h
        elif 'min capital' in h_lower or 'minimum capital' in h_lower or 'min investment' in h_lower:
            header_map['min_investment'] = h
        elif 'job creation' in h_lower or 'jobs req' in h_lower:
            header_map['job_creation_req'] = h
        elif 'incentive description' in h_lower or 'description' in h_lower:
            header_map['incentive_desc'] = h

    print(f"  Tax header map: {header_map}")

    d = {}
    for row in rows:
        s = row.get('State', '').strip()
        if s not in STATE_ABBREVS: continue
        d[s] = {
            'Tax Incentive?': row.get(header_map.get('tax_incentive', ''), '').strip(),
            'Incentive Type': row.get(header_map.get('incentive_type', ''), '').strip(),
            'Min Capital Investment': row.get(header_map.get('min_investment', ''), '').strip(),
            'Job Creation Req': row.get(header_map.get('job_creation_req', ''), '').strip(),
            'Incentive Description': row.get(header_map.get('incentive_desc', ''), '').strip(),
        }
    print(f"  Tax data: {len(d)} states")
    if d:
        sample = next(iter(d.items()))
        print(f"  Sample: {sample[0]} -> incentive='{sample[1]['Tax Incentive?']}'")
    return d

def read_regs(rows, headers):
    d = {}
    for row in rows:
        s = row.get('State', '').strip()
        if s not in STATE_ABBREVS: continue
        d[s] = safe_int(row.get('Total Regulations', ''))
    return d


# ═══════════════════════════════════════════
# EIA — Form EIA-861M aggregated sales/revenue file
# Contains monthly state×sector data from 1990 to present.
# Columns include Year, Month, State, Sector, plus revenue/sales/customers
# We compute price as (revenue × 100,000) / (sales_MWh × 1000) = cents/kWh
# ═══════════════════════════════════════════
def fetch_eia():
    """
    Download EIA-861M aggregated sales/revenue file with full monthly history.
    Returns: ({state: {(year, month): {'res': cents, 'com': cents, 'ind': cents}}}, latest_ym)
    """
    print("\n--- EIA-861M: Monthly sales/revenue (1990-present) ---")
    import openpyxl

    # Try several known URL patterns for the aggregated EIA-861M file
    candidate_urls = [
        "https://www.eia.gov/electricity/data/eia861m/archive/sales_revenue.xlsx",
        "https://www.eia.gov/electricity/data/eia861m/xls/sales_revenue.xlsx",
        "https://www.eia.gov/electricity/data/eia861m/sales_revenue.xlsx",
        # Fallback to the EPM 5.6.A table for current data only
        "https://www.eia.gov/electricity/monthly/xls/table_5_06_a.xlsx",
    ]

    for url in candidate_urls:
        print(f"  Trying: {url}")
        try:
            r = requests.get(url, timeout=120, allow_redirects=True,
                             headers={'User-Agent': 'Mozilla/5.0 datacenter-map-bot'})
            if r.status_code != 200:
                print(f"    Status {r.status_code}")
                continue

            ct = r.headers.get('content-type', '').lower()
            content_size = len(r.content)
            print(f"    Got {content_size:,} bytes, content-type: {ct}")

            # Heuristic: if it's smaller than 10KB it's probably an error page
            if content_size < 10000:
                print(f"    Too small, skipping")
                continue

            # Save and try to open
            with open('data/eia_raw.xlsx', 'wb') as f:
                f.write(r.content)

            try:
                wb = openpyxl.load_workbook('data/eia_raw.xlsx', data_only=True)
            except Exception as e:
                print(f"    Not a valid xlsx: {e}")
                continue

            print(f"    Sheets: {wb.sheetnames}")

            # Try each sheet, looking for one with year/month/state/sector columns
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row < 50:
                    continue

                # Find header row by scanning first 10 rows for "year" + "state"
                hrow = None
                for rr in range(1, 12):
                    row_vals = []
                    for cc in range(1, min(ws.max_column + 1, 30)):
                        v = ws.cell(row=rr, column=cc).value
                        row_vals.append(str(v or '').strip().lower())
                    has_year = any('year' == v for v in row_vals)
                    has_state = any('state' == v for v in row_vals)
                    if has_year and has_state:
                        hrow = rr
                        break

                if not hrow:
                    continue

                print(f"    Sheet '{sheet_name}': header row {hrow}")

                # Dump ALL header columns for diagnosis
                all_headers = {}
                for cc in range(1, ws.max_column + 1):
                    v = ws.cell(row=hrow, column=cc).value
                    if v is not None:
                        all_headers[cc] = str(v).strip()
                print(f"    All headers: {all_headers}")

                # Map columns
                cols = {}
                for cc, hdr in all_headers.items():
                    v = hdr.strip().lower()
                    if v == 'year': cols['year'] = cc
                    elif v == 'month': cols['month'] = cc
                    elif v == 'state': cols['state'] = cc
                    elif v == 'data status' or v == 'status':
                        pass  # skip
                    elif 'industry sector' in v or v == 'sector' or v == 'sectors' or 'sector category' in v:
                        cols['sector'] = cc
                    # Sales and revenue columns — names vary
                    elif 'residential' in v and 'sales' in v:
                        cols['res_sales'] = cc
                    elif 'residential' in v and 'revenue' in v:
                        cols['res_rev'] = cc
                    elif 'commercial' in v and 'sales' in v:
                        cols['com_sales'] = cc
                    elif 'commercial' in v and 'revenue' in v:
                        cols['com_rev'] = cc
                    elif 'industrial' in v and 'sales' in v:
                        cols['ind_sales'] = cc
                    elif 'industrial' in v and 'revenue' in v:
                        cols['ind_rev'] = cc
                    # Generic sales/revenue columns (long format)
                    elif v == 'sales' or 'megawatt' in v or 'mwh' in v:
                        if 'sales' not in cols: cols['sales'] = cc
                    elif v == 'revenue' or 'thousand dollars' in v or 'thousand_dollars' in v:
                        if 'revenue' not in cols: cols['revenue'] = cc

                print(f"    Mapped columns: {cols}")

                if 'year' not in cols or 'state' not in cols:
                    continue

                # Determine format: wide (separate cols per sector), long (sector column),
                # or simple (just total revenue+sales without sector breakdown)
                has_wide = 'res_sales' in cols and 'res_rev' in cols
                has_long = 'sector' in cols and 'sales' in cols and 'revenue' in cols
                has_simple = 'sales' in cols and 'revenue' in cols and not has_long

                # Skip YTD/annual sheets if we want monthly granularity
                is_monthly = 'month' in cols
                if not is_monthly:
                    print(f"    Skipping non-monthly sheet")
                    continue

                if not has_wide and not has_long and not has_simple:
                    print(f"    No usable sales/revenue columns found")
                    continue

                # Parse data rows
                result = {}  # state -> (year, month) -> {sec: cents/kWh}
                latest_ym = (1990, 1)
                row_count = 0

                for rr in range(hrow + 1, ws.max_row + 1):
                    yr = safe_int(ws.cell(row=rr, column=cols['year']).value)
                    if not yr or yr < 1990 or yr > 2050: continue

                    state_val = ws.cell(row=rr, column=cols['state']).value
                    if not state_val: continue
                    state_str = str(state_val).strip()

                    # Map abbreviation or name to canonical state name
                    if state_str in ABBREV_TO_STATE:
                        state_name = ABBREV_TO_STATE[state_str]
                    elif state_str in STATE_ABBREVS:
                        state_name = state_str
                    else:
                        continue

                    # Get month — may be numeric (1-12) or text
                    month_num = 1
                    if 'month' in cols:
                        m_val = ws.cell(row=rr, column=cols['month']).value
                        if m_val is not None:
                            try:
                                month_num = int(float(m_val))
                            except:
                                m_str = str(m_val).strip()[:3]
                                if m_str in MONTH_NAMES:
                                    month_num = MONTH_NAMES.index(m_str) + 1

                    ym = (yr, month_num)
                    if ym > latest_ym:
                        latest_ym = ym

                    if state_name not in result:
                        result[state_name] = {}
                    if ym not in result[state_name]:
                        result[state_name][ym] = {}

                    if has_wide:
                        # Wide format: separate res_sales/res_rev/com_sales/com_rev/ind_sales/ind_rev
                        for sec in ['res', 'com', 'ind']:
                            sales = safe_float(ws.cell(row=rr, column=cols[f'{sec}_sales']).value)
                            rev = safe_float(ws.cell(row=rr, column=cols[f'{sec}_rev']).value)
                            if sales and rev and sales > 0:
                                # revenue is in thousand dollars, sales in MWh
                                # cents/kWh = (rev_thousand_dollars × 100,000) / (sales_MWh × 1000)
                                #           = (rev × 100) / sales
                                cents_per_kwh = (rev * 100) / sales
                                if 0 < cents_per_kwh < 100:  # sanity check
                                    result[state_name][ym][sec] = cents_per_kwh
                                    row_count += 1

                    elif has_long:
                        # Long format: sector column + single sales/revenue
                        sec_str = str(ws.cell(row=rr, column=cols['sector']).value or '').strip().lower()
                        sales = safe_float(ws.cell(row=rr, column=cols['sales']).value)
                        rev = safe_float(ws.cell(row=rr, column=cols['revenue']).value)
                        if sales and rev and sales > 0:
                            cents_per_kwh = (rev * 100) / sales
                            if 0 < cents_per_kwh < 100:
                                if sec_str.startswith('res'):
                                    result[state_name][ym]['res'] = cents_per_kwh
                                    row_count += 1
                                elif sec_str.startswith('com'):
                                    result[state_name][ym]['com'] = cents_per_kwh
                                    row_count += 1
                                elif sec_str.startswith('ind'):
                                    result[state_name][ym]['ind'] = cents_per_kwh
                                    row_count += 1

                    elif has_simple:
                        # Simple format: total revenue + total sales, no sector breakdown
                        # Compute the all-sector average and assign to all three sectors
                        # (with a note in the data so the map can label it as "all-sector avg")
                        sales = safe_float(ws.cell(row=rr, column=cols['sales']).value)
                        rev = safe_float(ws.cell(row=rr, column=cols['revenue']).value)
                        if sales and rev and sales > 0:
                            cents_per_kwh = (rev * 100) / sales
                            if 0 < cents_per_kwh < 100:
                                # Assign the same all-sector value to all three sectors
                                # so the existing layer code works. We'll mark this as
                                # 'all-sector average' in the hover text via a flag.
                                result[state_name][ym]['res'] = cents_per_kwh
                                result[state_name][ym]['com'] = cents_per_kwh
                                result[state_name][ym]['ind'] = cents_per_kwh
                                result[state_name][ym]['_all_sector'] = True
                                row_count += 1

                if result and len(result) > 10:
                    print(f"    Parsed {len(result)} states, {row_count} state-month-sector combos")
                    print(f"    Latest: {MONTH_NAMES[latest_ym[1]-1]} {latest_ym[0]}")
                    sample = next(iter(result.keys()))
                    sample_data = result[sample].get(latest_ym, {})
                    print(f"    Sample [{sample}] {MONTH_NAMES[latest_ym[1]-1]} {latest_ym[0]}: {sample_data}")
                    if has_simple:
                        print(f"    NOTE: Using all-sector average prices (no sector breakdown in this file)")
                    return result, latest_ym

        except Exception as e:
            print(f"    Error: {e}")
            continue

    print("  EIA: No data retrieved from any candidate URL")
    return {}, None


# ═══════════════════════════════════════════
# BLS QCEW — fetch ALL years from latest-5 to latest
# ═══════════════════════════════════════════
def fetch_qcew_one(naics, year):
    url = f"https://data.bls.gov/cew/data/api/{year}/a/industry/{str(naics).replace('-','_')}.csv"
    try:
        r = requests.get(url, timeout=30)
        if r.status_code != 200: return None
    except: return None
    result = {}
    for row in csv.DictReader(io.StringIO(r.text)):
        af = row.get('area_fips','')
        oc = row.get('own_code','')
        if not af.endswith('000') or af == '00000' or oc not in ('0','5'):
            continue
        ab = FIPS_TO_ABBREV.get(af[:2])
        if not ab: continue
        sn = ABBREV_TO_STATE[ab]
        try:
            v = int(row.get('annual_avg_emplvl','0'))
            if sn not in result or v > result[sn]: result[sn] = v
        except: pass
    return result

def fetch_all_qcew():
    print("\n--- BLS QCEW employment ---")
    codes = {'dc_ops':'518210', 'comp_sys':'5415', 'nonres_build':'236220'}

    # Fetch every year from CURRENT_YEAR-7 to CURRENT_YEAR-1
    # so 1yr/3yr/5yr lookups always have a comparison year
    # (latest QCEW year is typically CURRENT_YEAR-1, so 5yr ago = CURRENT_YEAR-6)
    years_to_try = list(range(CURRENT_YEAR - 7, CURRENT_YEAR))

    out = {}
    for label, naics in codes.items():
        out[label] = {}
        for y in years_to_try:
            print(f"  {naics}/{y}", end='', flush=True)
            d = fetch_qcew_one(naics, y)
            if d:
                out[label][y] = d
                print(f" → {len(d)} states")
            else:
                print(" → no data")
            time.sleep(0.4)
    return out


# ═══════════════════════════════════════════
# CENSUS
# ═══════════════════════════════════════════
def fetch_census():
    print("\n--- Census population ---")
    for v in range(CURRENT_YEAR, CURRENT_YEAR - 4, -1):
        url = f"https://www2.census.gov/programs-surveys/popest/datasets/2020-{v}/state/totals/NST-EST{v}-ALLDATA.csv"
        print(f"  Trying vintage {v}...")
        try:
            r = requests.get(url, timeout=30)
            if r.status_code != 200: continue
            result = {}
            for row in csv.DictReader(io.StringIO(r.text)):
                s = row.get('NAME','').strip()
                if s not in STATE_ABBREVS: continue
                result[s] = {}
                for y in range(2020, v + 1):
                    val = safe_int(row.get(f'POPESTIMATE{y}',''))
                    if val: result[s][y] = val
            if result:
                print(f"  Vintage {v}: {len(result)} states")
                return result
        except: pass
    return {}


# ═══════════════════════════════════════════
# COMBINE
# ═══════════════════════════════════════════
def build_combined(dc_hist, lat_dc, tax, regs, qcew, census, eia, eia_ym, mkts, lat_mkt):
    census_yr = max(next(iter(census.values()), {}).keys(), default=None) if census else None

    # EIA latest year/month
    eia_label = ''
    if eia_ym:
        eia_label = f"{MONTH_NAMES[eia_ym[1]-1]} {eia_ym[0]}"

    state_list = []
    for state in sorted(dc_hist.keys()):
        cur_dcs = dc_hist[state].get(lat_dc)
        pop = census.get(state, {}).get(census_yr) if census_yr else None
        pc = round((cur_dcs/pop)*100000, 2) if cur_dcs and pop and pop > 0 else None

        e = {
            'state': state,
            'abbrev': STATE_ABBREVS.get(state, ''),
            'dcs': cur_dcs,
            'dcs_per_100k': pc,
            'population': pop,
            'census_year': census_yr,
            'dcs_month': f"{MONTH_NAMES[lat_dc[1]-1]} {lat_dc[0]}",
            'eia_month': eia_label,
            'tax_incentive': tax.get(state, {}).get('Tax Incentive?', ''),
            'incentive_type': tax.get(state, {}).get('Incentive Type', ''),
            'min_investment': tax.get(state, {}).get('Min Capital Investment', ''),
            'job_creation_req': tax.get(state, {}).get('Job Creation Req', ''),
            'incentive_desc': tax.get(state, {}).get('Incentive Description', ''),
            'total_regulations': regs.get(state),
        }

        # DC YoY % changes
        for w in CHANGE_WINDOWS:
            comp = dc_hist[state].get((lat_dc[0]-w, lat_dc[1]))
            e[f'dcs_{w}yr_ago'] = comp
            e[f'dcs_{w}yr_pct'] = pct_change(cur_dcs, comp)

        # Employment + % changes (latest year + 1/3/5 year ago)
        for lbl in qcew:
            yrs = sorted(qcew[lbl].keys(), reverse=True)
            if yrs:
                ly = yrs[0]
                cur_emp = qcew[lbl].get(ly, {}).get(state)
                e[f'{lbl}_current'] = cur_emp
                e[f'{lbl}_year'] = ly
                for w in CHANGE_WINDOWS:
                    h = qcew[lbl].get(ly - w, {}).get(state)
                    e[f'{lbl}_{w}yr_pct'] = pct_change(cur_emp, h)
                    e[f'{lbl}_{w}yr_year'] = ly - w if h else None

        # Population % changes
        if census_yr and state in census:
            for w in CHANGE_WINDOWS:
                h = census[state].get(census_yr - w)
                e[f'pop_{w}yr_pct'] = pct_change(pop, h)

        # Electricity rates + % changes (1yr/3yr/5yr)
        # eia[state] is a dict of (year, month) → {'res':, 'com':, 'ind':}
        # Now sourced from EIA-861M with full monthly history back to 2010+
        if eia_ym and state in eia and eia_ym in eia[state]:
            current_rates = eia[state][eia_ym]
            for sec in ['res', 'com', 'ind']:
                cur_rate = current_rates.get(sec)
                e[f'{sec}_rate'] = cur_rate
                # For each window, look up the same month N years prior
                for w in CHANGE_WINDOWS:
                    ym_hist = (eia_ym[0] - w, eia_ym[1])
                    hist_rate = eia.get(state, {}).get(ym_hist, {}).get(sec)
                    e[f'{sec}_{w}yr_pct'] = pct_change(cur_rate, hist_rate)
                    e[f'{sec}_{w}yr_ago'] = hist_rate

        state_list.append(e)

    # Rankings
    state_list.sort(key=lambda x: x.get('dcs') or 0, reverse=True)
    for i, s in enumerate(state_list, 1): s['rank_total'] = i
    by_pc = sorted(state_list, key=lambda x: x.get('dcs_per_100k') or 0, reverse=True)
    for i, s in enumerate(by_pc, 1): s['rank_per_capita'] = i

    # Markets
    mkt_list = []
    for (state, city), hist in mkts.items():
        cur = hist.get(lat_mkt)
        m = {
            'state': state, 'city': city, 'dcs': cur,
            'month': f"{MONTH_NAMES[lat_mkt[1]-1]} {lat_mkt[0]}",
        }
        for w in CHANGE_WINDOWS:
            comp = hist.get((lat_mkt[0]-w, lat_mkt[1]))
            m[f'dcs_{w}yr_ago'] = comp
            m[f'dcs_{w}yr_pct'] = pct_change(cur, comp)
        mkt_list.append(m)

    out = {
        'generated': datetime.now().isoformat(),
        'sheet_id': GOOGLE_SHEET_ID,
        'latest_dc_month': f"{MONTH_NAMES[lat_dc[1]-1]} {lat_dc[0]}",
        'latest_mkt_month': f"{MONTH_NAMES[lat_mkt[1]-1]} {lat_mkt[0]}",
        'latest_eia_month': eia_label,
        'change_windows': CHANGE_WINDOWS,
        'states': state_list,
        'markets': mkt_list,
    }
    with open('data/combined.json','w') as f:
        json.dump(out, f, indent=2, default=str)
    print(f"\n  Output: {len(state_list)} states, {len(mkt_list)} markets")
    return out


# ═══════════════════════════════════════════
def main():
    print("="*60)
    print("USA Data Center Map — Data Pipeline")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("="*60)
    os.makedirs('data', exist_ok=True)

    print("\n--- Google Sheet ---")
    dc_rows, dc_h = fetch_sheet('State DC History')
    dc_hist, lat_dc = read_dc_history(dc_rows, dc_h)
    tax_rows, tax_h = fetch_sheet('State Tax Incentive', debug=True)
    tax = read_tax(tax_rows, tax_h)
    reg_rows, reg_h = fetch_sheet('Total Regulations')
    regs = read_regs(reg_rows, reg_h)
    mkt_rows, mkt_h = fetch_sheet('Regional Market History')
    mkts, lat_mkt = read_markets(mkt_rows, mkt_h)

    eia, eia_ym = fetch_eia()
    qcew = fetch_all_qcew()
    census = fetch_census()

    print("\n--- Building combined output ---")
    build_combined(dc_hist, lat_dc, tax, regs, qcew, census, eia, eia_ym, mkts, lat_mkt)

    with open('data/last_updated.json','w') as f:
        json.dump({
            'timestamp': datetime.now().isoformat(),
            'sheet_id': GOOGLE_SHEET_ID,
        }, f, indent=2)
    print("\n" + "="*60 + "\nDONE\n" + "="*60)


if __name__ == '__main__':
    main()
